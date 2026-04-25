#!/usr/bin/env python3
"""
Extrae archivos CUIPO per-municipality desde las bases consolidadas descargadas
de CIFFIT/DNP (ciffit.dnp.gov.co) y genera fixtures .xlsx por municipio en el
formato que los parsers de chip-parser.ts esperan.

Archivos de entrada (consolidados — contienen TODOS los municipios):
  *CUIPOA_PROGRAMACION_DE_INGRESOS.xlsx  (header row 16, 0-indexed)
  *CUIPOB_EJECUCION_DE_INGRESOS.xlsb     (header row 16)
  *CUIPOC_PROGRAMACION_DE_GASTOS.xlsb     (header row 16)
  *CUIPOD_EJECUCION_DE_GASTOS.xlsb        (header row 17)

Archivos de salida (por municipio, formato CHIP per-entity):
  {out_dir}/{dane}/cuipo_prog_ing.xlsx
  {out_dir}/{dane}/cuipo_ejec_ing.xlsx
  {out_dir}/{dane}/cuipo_prog_gas.xlsx
  {out_dir}/{dane}/cuipo_ejec_gas.xlsx

Uso:
  # Extraer todos los municipios de Antioquia (depto 05)
  python3 scripts/extract-cuipo-fixtures.py ~/Downloads/ public/fixtures/

  # Extraer un municipio específico
  python3 scripts/extract-cuipo-fixtures.py ~/Downloads/ public/fixtures/ --dane 05091

  # Extraer varios municipios
  python3 scripts/extract-cuipo-fixtures.py ~/Downloads/ public/fixtures/ --dane 05091 --dane 05001

  # Extraer un departamento distinto
  python3 scripts/extract-cuipo-fixtures.py ~/Downloads/ public/fixtures/ --dept 25
"""

import argparse
import os
import re
import sys
import time
from collections import defaultdict
from pathlib import Path
from typing import Optional

# ---------------------------------------------------------------------------
# Column layouts for consolidated DB files (0-indexed column numbers)
# All files have 1-based column offsets (col 0 is always empty/None)
# ---------------------------------------------------------------------------

# CUIPOA — Programación de Ingresos (.xlsx, header row 16)
A_HEADER_ROW = 16
A_COL_DANE = 4
A_COL_FUT = 5
A_COL_ENTIDAD = 6
A_COL_CONCEPTO_COD = 11
A_COL_CONCEPTO = 12
A_COL_SECTORIAL_COD = 13
A_COL_SECTORIAL = 14
A_COL_PPTO_INICIAL = 15
A_COL_PPTO_DEFINITIVO = 16

# CUIPOB — Ejecución de Ingresos (.xlsb, header row 16)
B_HEADER_ROW = 16
B_COL_DANE = 4
B_COL_FUT = 5
B_COL_ENTIDAD = 6
B_COL_CONCEPTO_COD = 11
B_COL_CONCEPTO = 12
B_COL_CPC_COD = 13
B_COL_CPC = 14
B_COL_DET_SECT_COD = 15
B_COL_DET_SECTORIAL = 16
B_COL_FUENTE_COD = 17
B_COL_FUENTE = 18
B_COL_TERCERO_COD = 19
B_COL_TERCERO = 20
B_COL_POL_PUB_COD = 21
B_COL_POL_PUBLICA = 22
B_COL_NORMA = 23
B_COL_TIPO_NORMA_COD = 24
B_COL_TIPO_NORMA = 25
B_COL_REC_VA_SIN = 26
B_COL_REC_VA_CON = 27
B_COL_REC_ACTUAL_SIN = 28
B_COL_REC_ACTUAL_CON = 29
B_COL_TOTAL_RECAUDO = 30

# CUIPOC — Programación de Gastos (.xlsb, header row 16)
C_HEADER_ROW = 16
C_COL_DANE = 4
C_COL_FUT = 5
C_COL_ENTIDAD = 6
C_COL_CONCEPTO_COD = 11
C_COL_CONCEPTO = 12
C_COL_VIG_GASTO_COD = 13
C_COL_VIG_GASTO = 14
C_COL_SECCION_COD = 15
C_COL_SECCION = 16
C_COL_PROG_MGA_COD = 17
C_COL_PROG_MGA = 18
C_COL_BPIN = 19
C_COL_APROP_INICIAL = 20
C_COL_APROP_DEFINITIVA = 21

# CUIPOD — Ejecución de Gastos (.xlsb, header row 17)
D_HEADER_ROW = 17
D_COL_DANE = 4
D_COL_FUT = 5
D_COL_ENTIDAD = 6
D_COL_CONCEPTO_COD = 11
D_COL_CONCEPTO = 12
D_COL_VIG_GASTO_COD = 13
D_COL_VIG_GASTO = 14
D_COL_SECCION_COD = 15
D_COL_SECCION = 16
D_COL_PROD_MGA_COD = 17
D_COL_PROD_MGA = 18
D_COL_CPC_COD = 19
D_COL_CPC = 20
D_COL_DET_SECT_COD = 21
D_COL_DET_SECTORIAL = 22
D_COL_FUENTE_COD = 23
D_COL_FUENTE = 24
D_COL_BPIN = 25
D_COL_SIT_FONDOS_COD = 26
D_COL_SIT_FONDOS = 27
D_COL_POL_PUB_COD = 28
D_COL_POL_PUBLICA = 29
D_COL_TERCERO_COD = 30
D_COL_TERCERO = 31
D_COL_COMPROMISOS = 32
D_COL_OBLIGACIONES = 33
D_COL_PAGOS = 34

MAX_COL = 40


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def sv(val) -> str:
    """String value, stripped."""
    if val is None:
        return ""
    s = str(val).strip()
    return s


def nv(val) -> str:
    """Numeric value as string for output cells (trailing space for CHIP format)."""
    if val is None:
        return "0 "
    if isinstance(val, (int, float)):
        # Integer if no decimals
        if val == int(val):
            return f"{int(val)} "
        return f"{val} "
    s = str(val).strip()
    if not s:
        return "0 "
    try:
        f = float(s)
        if f == int(f):
            return f"{int(f)} "
        return f"{f} "
    except ValueError:
        return "0 "


def detect_files(input_dir: str) -> dict:
    """Auto-detect which consolidated CUIPO files exist in the given directory.

    Returns dict with keys 'A', 'B', 'C', 'D' mapping to file paths.
    """
    found = {}
    p = Path(input_dir)
    for f in p.iterdir():
        if f.name.startswith("~$"):
            continue
        name_upper = f.name.upper()
        if "CUIPOA" in name_upper and f.suffix.lower() in (".xlsx", ".xlsb"):
            found["A"] = str(f)
        elif "CUIPOB" in name_upper and f.suffix.lower() in (".xlsx", ".xlsb"):
            found["B"] = str(f)
        elif "CUIPOC" in name_upper and f.suffix.lower() in (".xlsx", ".xlsb"):
            found["C"] = str(f)
        elif "CUIPOD" in name_upper and f.suffix.lower() in (".xlsx", ".xlsb"):
            found["D"] = str(f)
    return found


def read_xlsb_rows(path: str, header_row: int) -> list[list]:
    """Read .xlsb file using pyxlsb. Returns list of rows (each row = list of values)."""
    from pyxlsb import open_workbook

    rows = []
    with open_workbook(path) as wb:
        sheet = wb.sheets[0]
        with wb.get_sheet(sheet) as sh:
            for i, row in enumerate(sh.rows()):
                if i <= header_row:
                    continue  # skip header and preamble
                arr = [None] * (MAX_COL + 1)
                has = False
                for cell in row:
                    if cell.c <= MAX_COL and cell.v is not None:
                        v = cell.v
                        if isinstance(v, str):
                            v = v.strip()
                        arr[cell.c] = v
                        has = True
                if has:
                    rows.append(arr)
    return rows


def read_xlsx_rows(path: str, header_row: int) -> list[list]:
    """Read .xlsx file using openpyxl. Returns list of rows."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_row:
            continue
        arr = [None] * (MAX_COL + 1)
        has = False
        for c, v in enumerate(row):
            if c > MAX_COL:
                break
            if v is not None:
                if isinstance(v, str):
                    v = v.strip()
                arr[c] = v
                has = True
        if has:
            rows.append(arr)
    wb.close()
    return rows


def read_rows(path: str, header_row: int) -> list[list]:
    """Read rows from either .xlsx or .xlsb file."""
    if path.lower().endswith(".xlsb"):
        return read_xlsb_rows(path, header_row)
    else:
        return read_xlsx_rows(path, header_row)


def group_by_dane(rows: list[list], dane_col: int) -> dict[str, list[list]]:
    """Group rows by DANE code. Returns {dane: [rows]}."""
    groups = defaultdict(list)
    for row in rows:
        dane = sv(row[dane_col])
        if not dane:
            continue
        # Normalize DANE code to 5 digits
        dane = dane.zfill(5)
        groups[dane].append(row)
    return dict(groups)


# ---------------------------------------------------------------------------
# CHIP-format output writers
# Each writes an .xlsx matching the per-entity CHIP download format.
# ---------------------------------------------------------------------------

def write_chip_preamble(ws, chip_code: str, entity_name: str,
                        section_label: str, n_cols: int):
    """Write the standard CHIP preamble rows (rows 1-10 in the output).

    Format:
      Row 1-2: blank
      Row 3: CHIP code + entity name
      Row 4: "MUNICIPIOS"
      Row 5: date range
      Row 6: "CUIPO - CATEGORIA UNICA ..."
      Row 7: section label (A_PROGRAMACION_DE_INGRESOS, etc.)
      Row 8: "ENVIO NUMERO ..."
      Row 9: "FECHA RECEPCION ..."
      Row 10-11: blank
      Row 12: HEADER
    """
    # Rows 1-2: blank
    ws.append([None] * n_cols)
    ws.append([None] * n_cols)

    # Row 3: entity
    row3 = [None] * n_cols
    row3[0] = f"{chip_code} - {entity_name} "
    ws.append(row3)

    # Row 4: MUNICIPIOS
    row4 = [None] * n_cols
    row4[0] = "MUNICIPIOS "
    ws.append(row4)

    # Row 5: date range (use generic current period)
    row5 = [None] * n_cols
    row5[0] = "01-12-2025 al 31-12-2025"
    ws.append(row5)

    # Row 6: CUIPO title
    row6 = [None] * n_cols
    row6[0] = "CUIPO - CATEGORIA UNICA DE INFORMACION DEL PRESUPUESTO ORDINARIO "
    ws.append(row6)

    # Row 7: section
    row7 = [None] * n_cols
    row7[0] = f"{section_label} "
    ws.append(row7)

    # Row 8: envio
    row8 = [None] * n_cols
    row8[0] = "ENVIO NUMERO -"
    ws.append(row8)

    # Row 9: fecha
    row9 = [None] * n_cols
    row9[0] = "FECHA RECEPCION -"
    ws.append(row9)

    # Rows 10-11: blank
    ws.append([None] * n_cols)
    ws.append([None] * n_cols)


def write_prog_ing(dane: str, rows: list[list], entity_name: str,
                   chip_code: str, output_dir: str):
    """Write cuipo_prog_ing.xlsx — Programación de Ingresos."""
    import openpyxl

    headers = [
        "CODIGO",
        "NOMBRE",
        "DETALLE SECTORIAL",
        "PRESUPUESTO INICIAL(Pesos)",
        "PRESUPUESTO DEFINITIVO(Pesos)",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PROG_ING"

    write_chip_preamble(ws, chip_code, entity_name,
                        "A_PROGRAMACION_DE_INGRESOS", len(headers))

    # Header row (will be row 12 in Excel, index 11 0-based)
    ws.append(headers)

    # Data rows
    for row in rows:
        concepto = sv(row[A_COL_CONCEPTO_COD])
        if not concepto:
            continue
        ws.append([
            concepto + " ",
            sv(row[A_COL_CONCEPTO]) + " ",
            sv(row[A_COL_SECTORIAL]) + " " if sv(row[A_COL_SECTORIAL]) else "  ",
            nv(row[A_COL_PPTO_INICIAL]),
            nv(row[A_COL_PPTO_DEFINITIVO]),
        ])

    out_path = os.path.join(output_dir, dane, "cuipo_prog_ing.xlsx")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return len(rows)


def write_ejec_ing(dane: str, rows: list[list], entity_name: str,
                   chip_code: str, output_dir: str):
    """Write cuipo_ejec_ing.xlsx — Ejecución de Ingresos."""
    import openpyxl

    headers = [
        "CODIGO",
        "NOMBRE",
        "CPC",
        "DETALLE SECTORIAL",
        "FUENTES DE FINANCIACION",
        "TERCEROS",
        "POLITICA PUBLICA",
        "NUMERO Y FECHA DE LA NORMA",
        "TIPO DE NORMA",
        "RECAUDO VIGEN ACTUAL SIN FONDOS(Pesos)",
        "RECAUDO VIGEN ACTUAL CON FONDOS(Pesos)",
        "RECAUDO VIGEN ANTERIOR SIN FONDO(Pesos)",
        "RECAUDO VIGEN ANTERIOR CON FONDO(Pesos)",
        "TOTAL RECAUDO(Pesos)",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EJEC_ING"

    write_chip_preamble(ws, chip_code, entity_name,
                        "B_EJECUCION_DE_INGRESOS", len(headers))

    ws.append(headers)

    for row in rows:
        concepto = sv(row[B_COL_CONCEPTO_COD])
        if not concepto:
            continue
        ws.append([
            concepto + " ",
            sv(row[B_COL_CONCEPTO]) + " ",
            sv(row[B_COL_CPC]) + " " if sv(row[B_COL_CPC]) else "  ",
            sv(row[B_COL_DET_SECTORIAL]) + " " if sv(row[B_COL_DET_SECTORIAL]) else "  ",
            sv(row[B_COL_FUENTE]) + " " if sv(row[B_COL_FUENTE]) else "  ",
            sv(row[B_COL_TERCERO]) + " " if sv(row[B_COL_TERCERO]) else "  ",
            sv(row[B_COL_POL_PUBLICA]) + " " if sv(row[B_COL_POL_PUBLICA]) else "  ",
            sv(row[B_COL_NORMA]) + " " if sv(row[B_COL_NORMA]) else "  ",
            sv(row[B_COL_TIPO_NORMA]) + " " if sv(row[B_COL_TIPO_NORMA]) else "  ",
            nv(row[B_COL_REC_ACTUAL_SIN]),
            nv(row[B_COL_REC_ACTUAL_CON]),
            nv(row[B_COL_REC_VA_SIN]),
            nv(row[B_COL_REC_VA_CON]),
            nv(row[B_COL_TOTAL_RECAUDO]),
        ])

    out_path = os.path.join(output_dir, dane, "cuipo_ejec_ing.xlsx")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return len(rows)


def write_prog_gas(dane: str, rows: list[list], entity_name: str,
                   chip_code: str, output_dir: str):
    """Write cuipo_prog_gas.xlsx — Programación de Gastos."""
    import openpyxl

    headers = [
        "VIGENCIA GASTO",
        "SECCION PRESUPUESTAL",
        "CODIGO",
        "NOMBRE",
        "PROGRAMA MGA",
        "DETALLE SECTORIAL",
        "BPIN",
        "APROPIACION INICIAL(Pesos)",
        "APROPIACION DEFINITIVA(Pesos)",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PROG_GAS"

    write_chip_preamble(ws, chip_code, entity_name,
                        "C_PROGRAMACION_DE_GASTOS", len(headers))

    ws.append(headers)

    # Track last vigencia/seccion for fill-down behavior in output
    last_vig = ""
    last_sec = ""

    for row in rows:
        concepto = sv(row[C_COL_CONCEPTO_COD])
        if not concepto:
            continue

        vig = sv(row[C_COL_VIG_GASTO])
        sec = sv(row[C_COL_SECCION])

        # Fill-down: only show vigencia/seccion when they change
        show_vig = ""
        show_sec = ""
        if vig and vig != last_vig:
            show_vig = vig + " "
            last_vig = vig
        if sec and sec != last_sec:
            show_sec = sec + " "
            last_sec = sec

        # Use \xa0 (non-breaking space) for empty vigencia/seccion cells
        # to match the CHIP format
        ws.append([
            show_vig if show_vig else "\xa0",
            show_sec if show_sec else "\xa0",
            concepto + " ",
            sv(row[C_COL_CONCEPTO]) + " ",
            sv(row[C_COL_PROG_MGA]) + " " if sv(row[C_COL_PROG_MGA]) else "  ",
            "  ",  # Detalle Sectorial (not in consolidated C)
            sv(row[C_COL_BPIN]) + " " if sv(row[C_COL_BPIN]) else "  ",
            nv(row[C_COL_APROP_INICIAL]),
            nv(row[C_COL_APROP_DEFINITIVA]),
        ])

    out_path = os.path.join(output_dir, dane, "cuipo_prog_gas.xlsx")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return len(rows)


def write_ejec_gas(dane: str, rows: list[list], entity_name: str,
                   chip_code: str, output_dir: str):
    """Write cuipo_ejec_gas.xlsx — Ejecución de Gastos."""
    import openpyxl

    headers = [
        "VIGENCIA GASTO",
        "SECCION PRESUPUESTAL",
        "CODIGO",
        "NOMBRE",
        "PRODUCTO MGA",
        "CPC",
        "DETALLE SECTORIAL",
        "FUENTES DE FINANCIACION",
        "BPIN",
        "SITUACION DE FONDOS",
        "POLITICA PUBLICA",
        "TERCEROS",
        "COMPROMISOS(Pesos)",
        "OBLIGACIONES(Pesos)",
        "PAGOS(Pesos)",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EJEC_GAS"

    write_chip_preamble(ws, chip_code, entity_name,
                        "D_EJECUCION_DE_GASTOS", len(headers))

    ws.append(headers)

    last_vig = ""
    last_sec = ""

    for row in rows:
        concepto = sv(row[D_COL_CONCEPTO_COD])
        if not concepto:
            continue

        vig = sv(row[D_COL_VIG_GASTO])
        sec = sv(row[D_COL_SECCION])

        show_vig = ""
        show_sec = ""
        if vig and vig != last_vig:
            show_vig = vig + " "
            last_vig = vig
        if sec and sec != last_sec:
            show_sec = sec + " "
            last_sec = sec

        ws.append([
            show_vig if show_vig else "\xa0",
            show_sec if show_sec else "\xa0",
            concepto + " ",
            sv(row[D_COL_CONCEPTO]) + " ",
            sv(row[D_COL_PROD_MGA]) + " " if sv(row[D_COL_PROD_MGA]) else "  ",
            sv(row[D_COL_CPC]) + " " if sv(row[D_COL_CPC]) else "  ",
            sv(row[D_COL_DET_SECTORIAL]) + " " if sv(row[D_COL_DET_SECTORIAL]) else "  ",
            sv(row[D_COL_FUENTE]) + " " if sv(row[D_COL_FUENTE]) else "  ",
            sv(row[D_COL_BPIN]) + " " if sv(row[D_COL_BPIN]) else "  ",
            sv(row[D_COL_SIT_FONDOS]) + " " if sv(row[D_COL_SIT_FONDOS]) else "  ",
            sv(row[D_COL_POL_PUBLICA]) + " " if sv(row[D_COL_POL_PUBLICA]) else "  ",
            sv(row[D_COL_TERCERO]) + " " if sv(row[D_COL_TERCERO]) else "  ",
            nv(row[D_COL_COMPROMISOS]),
            nv(row[D_COL_OBLIGACIONES]),
            nv(row[D_COL_PAGOS]),
        ])

    out_path = os.path.join(output_dir, dane, "cuipo_ejec_gas.xlsx")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return len(rows)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def build_chip_code(dane: str, fut_code: str) -> str:
    """Build a CHIP-style entity code from DANE and FUT codes.

    CHIP codes look like '219105091' — the FUT code from the consolidated file.
    If FUT is not available, synthesize from DANE.
    """
    fut = sv(fut_code)
    if fut:
        return fut
    # Fallback: typical municipal FUT = "2191" + dane5
    return f"2191{dane.zfill(5)}"


def main():
    parser = argparse.ArgumentParser(
        description="Extract per-municipality CUIPO fixtures from consolidated DB files"
    )
    parser.add_argument(
        "input_dir",
        help="Directory containing the 4 consolidated CUIPO files"
    )
    parser.add_argument(
        "output_dir",
        help="Output directory for per-municipality fixtures"
    )
    parser.add_argument(
        "--dane",
        action="append",
        dest="dane_codes",
        help="Specific DANE code(s) to extract (5-digit). Can be repeated."
    )
    parser.add_argument(
        "--dept",
        default="05",
        help="Department code prefix to filter (default: 05 = Antioquia)"
    )
    parser.add_argument(
        "--all",
        action="store_true",
        dest="extract_all",
        help="Extract ALL municipalities, ignoring department filter"
    )

    args = parser.parse_args()

    input_dir = os.path.expanduser(args.input_dir)
    output_dir = os.path.expanduser(args.output_dir)

    if not os.path.isdir(input_dir):
        print(f"ERROR: Input directory not found: {input_dir}")
        return 1

    # Detect files
    files = detect_files(input_dir)
    found_labels = {
        "A": "Prog. Ingresos",
        "B": "Ejec. Ingresos",
        "C": "Prog. Gastos",
        "D": "Ejec. Gastos",
    }

    print("=== CUIPO Fixture Extractor ===")
    print(f"Input:  {input_dir}")
    print(f"Output: {output_dir}")
    print()

    for kind in ["A", "B", "C", "D"]:
        if kind in files:
            print(f"  [{kind}] {found_labels[kind]}: {os.path.basename(files[kind])}")
        else:
            print(f"  [{kind}] {found_labels[kind]}: NOT FOUND")

    if not files:
        print("\nERROR: No CUIPO files found in the input directory.")
        return 1

    print()

    # Determine which DANE codes to extract
    target_danes: Optional[set] = None
    if args.dane_codes:
        target_danes = {d.zfill(5) for d in args.dane_codes}
        print(f"Filter: specific DANEs = {sorted(target_danes)}")
    elif not args.extract_all:
        dept = args.dept
        print(f"Filter: department prefix = {dept}")
    else:
        print("Filter: ALL municipalities")

    def should_include(dane: str) -> bool:
        if target_danes is not None:
            return dane in target_danes
        if args.extract_all:
            return True
        return dane.startswith(args.dept)

    # ---------------------------------------------------------------------------
    # Process each file type
    # ---------------------------------------------------------------------------

    # Collect entity metadata (name, FUT code) from first file we encounter
    entity_meta: dict[str, tuple[str, str]] = {}  # dane -> (name, fut_code)

    # Track stats
    stats = {
        "A": {"munis": 0, "rows": 0},
        "B": {"munis": 0, "rows": 0},
        "C": {"munis": 0, "rows": 0},
        "D": {"munis": 0, "rows": 0},
    }

    # ---- CUIPOA: Programación de Ingresos ----
    if "A" in files:
        print(f"\n--- Processing CUIPOA (Prog. Ingresos) ---")
        t0 = time.time()
        all_rows = read_rows(files["A"], A_HEADER_ROW)
        dt = time.time() - t0
        print(f"  Read {len(all_rows):,} rows in {dt:.1f}s")

        grouped = group_by_dane(all_rows, A_COL_DANE)
        for dane, rows in sorted(grouped.items()):
            if not should_include(dane):
                continue
            name = sv(rows[0][A_COL_ENTIDAD])
            fut = sv(rows[0][A_COL_FUT])
            entity_meta[dane] = (name, fut)
            chip_code = build_chip_code(dane, fut)
            n = write_prog_ing(dane, rows, name, chip_code, output_dir)
            stats["A"]["munis"] += 1
            stats["A"]["rows"] += n
            print(f"  {dane} ({name[:30]}): {n} rows")

        del all_rows, grouped

    # ---- CUIPOB: Ejecución de Ingresos ----
    if "B" in files:
        print(f"\n--- Processing CUIPOB (Ejec. Ingresos) ---")
        t0 = time.time()
        all_rows = read_rows(files["B"], B_HEADER_ROW)
        dt = time.time() - t0
        print(f"  Read {len(all_rows):,} rows in {dt:.1f}s")

        grouped = group_by_dane(all_rows, B_COL_DANE)
        for dane, rows in sorted(grouped.items()):
            if not should_include(dane):
                continue
            name = sv(rows[0][B_COL_ENTIDAD])
            fut = sv(rows[0][B_COL_FUT])
            if dane not in entity_meta:
                entity_meta[dane] = (name, fut)
            chip_code = build_chip_code(dane, entity_meta.get(dane, (name, fut))[1])
            n = write_ejec_ing(dane, rows, name, chip_code, output_dir)
            stats["B"]["munis"] += 1
            stats["B"]["rows"] += n
            print(f"  {dane} ({name[:30]}): {n} rows")

        del all_rows, grouped

    # ---- CUIPOC: Programación de Gastos ----
    if "C" in files:
        print(f"\n--- Processing CUIPOC (Prog. Gastos) ---")
        t0 = time.time()
        all_rows = read_rows(files["C"], C_HEADER_ROW)
        dt = time.time() - t0
        print(f"  Read {len(all_rows):,} rows in {dt:.1f}s")

        grouped = group_by_dane(all_rows, C_COL_DANE)
        for dane, rows in sorted(grouped.items()):
            if not should_include(dane):
                continue
            name = sv(rows[0][C_COL_ENTIDAD])
            fut = sv(rows[0][C_COL_FUT])
            if dane not in entity_meta:
                entity_meta[dane] = (name, fut)
            chip_code = build_chip_code(dane, entity_meta.get(dane, (name, fut))[1])
            n = write_prog_gas(dane, rows, name, chip_code, output_dir)
            stats["C"]["munis"] += 1
            stats["C"]["rows"] += n
            print(f"  {dane} ({name[:30]}): {n} rows")

        del all_rows, grouped

    # ---- CUIPOD: Ejecución de Gastos ----
    if "D" in files:
        print(f"\n--- Processing CUIPOD (Ejec. Gastos) ---")
        t0 = time.time()
        all_rows = read_rows(files["D"], D_HEADER_ROW)
        dt = time.time() - t0
        print(f"  Read {len(all_rows):,} rows in {dt:.1f}s")

        grouped = group_by_dane(all_rows, D_COL_DANE)
        for dane, rows in sorted(grouped.items()):
            if not should_include(dane):
                continue
            name = sv(rows[0][D_COL_ENTIDAD])
            fut = sv(rows[0][D_COL_FUT])
            if dane not in entity_meta:
                entity_meta[dane] = (name, fut)
            chip_code = build_chip_code(dane, entity_meta.get(dane, (name, fut))[1])
            n = write_ejec_gas(dane, rows, name, chip_code, output_dir)
            stats["D"]["munis"] += 1
            stats["D"]["rows"] += n
            print(f"  {dane} ({name[:30]}): {n} rows")

        del all_rows, grouped

    # Summary
    print("\n=== Summary ===")
    total_munis = len(entity_meta)
    for kind in ["A", "B", "C", "D"]:
        s = stats[kind]
        if s["munis"] > 0:
            print(f"  [{kind}] {found_labels[kind]}: {s['munis']} municipalities, {s['rows']:,} rows")

    all_danes = set()
    for kind in ["A", "B", "C", "D"]:
        if kind in files:
            # We already know which danes were processed from stats
            pass

    print(f"\n  Total municipalities with at least 1 fixture: {total_munis}")
    print(f"  Output directory: {output_dir}")
    print("  Done!")
    return 0


if __name__ == "__main__":
    sys.exit(main())
