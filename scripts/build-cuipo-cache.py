#!/usr/bin/env python3
"""
Pre-procesa archivos CUIPO .xlsb / .xlsx descargados a JSON cache.

pyxlsb es ~50x más rápido que SheetJS (Node) para parsear .xlsb. Este script
convierte cada archivo a un JSON con las columnas usadas por
src/lib/cuipo-local-xlsb.ts y lo deja en ~/.cache/gobia-cuipo/.

Uso:
  python3 scripts/build-cuipo-cache.py [archivo1.xlsb ...]
  (sin args: procesa todos los CUIPO* en ~/Downloads)
"""
import json
import os
import sys
import time
from pathlib import Path

CACHE_DIR = Path.home() / ".cache" / "gobia-cuipo"
DOWNLOADS = Path.home() / "Downloads"
MAX_COL = 40


def is_xlsb(p: Path) -> bool:
    return p.suffix.lower() == ".xlsb"


def is_xlsx(p: Path) -> bool:
    return p.suffix.lower() == ".xlsx"


def cache_path_for(src: Path) -> Path:
    return CACHE_DIR / (src.stem + ".json")


def needs_rebuild(src: Path, dst: Path) -> bool:
    if not dst.exists():
        return True
    return dst.stat().st_mtime < src.stat().st_mtime


def convert_xlsb(src: Path, dst: Path) -> int:
    from pyxlsb import open_workbook
    rows_out = []
    with open_workbook(str(src)) as wb:
        sheet = wb.sheets[0]
        with wb.get_sheet(sheet) as sh:
            for row in sh.rows():
                arr = [None] * (MAX_COL + 1)
                has = False
                for cell in row:
                    if cell.c <= MAX_COL and cell.v is not None:
                        v = cell.v
                        if isinstance(v, str):
                            v = v.strip() or None
                        if v is not None:
                            arr[cell.c] = v
                            has = True
                if has:
                    rows_out.append(arr)
    dst.parent.mkdir(parents=True, exist_ok=True)
    with open(dst, "w") as f:
        json.dump(rows_out, f, separators=(",", ":"), default=str)
    return len(rows_out)


def convert_xlsx(src: Path, dst: Path) -> int:
    import openpyxl
    wb = openpyxl.load_workbook(str(src), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_out = []
    for row in ws.iter_rows(values_only=True):
        arr = [None] * (MAX_COL + 1)
        has = False
        for c, v in enumerate(row):
            if c > MAX_COL:
                break
            if v is None:
                continue
            if isinstance(v, str):
                v = v.strip() or None
            if v is None:
                continue
            arr[c] = v
            has = True
        if has:
            rows_out.append(arr)
    wb.close()
    dst.parent.mkdir(parents=True, exist_ok=True)
    with open(dst, "w") as f:
        json.dump(rows_out, f, separators=(",", ":"), default=str)
    return len(rows_out)


def main():
    if len(sys.argv) > 1:
        files = [Path(p) for p in sys.argv[1:]]
    else:
        files = sorted(DOWNLOADS.glob("*_CUIPO*"))
        # Excluye archivos temporales de Excel (~$...) y cualquier otro patrón raro
        files = [
            f for f in files
            if (is_xlsb(f) or is_xlsx(f))
            and not f.name.startswith("~$")
        ]
    if not files:
        print("No hay archivos CUIPO* en", DOWNLOADS)
        return 1
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    for src in files:
        dst = cache_path_for(src)
        if not needs_rebuild(src, dst):
            print(f"  cache fresh:   {dst.name}")
            continue
        t0 = time.time()
        if is_xlsb(src):
            n = convert_xlsb(src, dst)
        elif is_xlsx(src):
            n = convert_xlsx(src, dst)
        else:
            continue
        dt = time.time() - t0
        size_mb = dst.stat().st_size / 1024 / 1024
        print(f"  built:         {dst.name}  ({n:,} rows, {size_mb:.1f}MB, {dt:.1f}s)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
